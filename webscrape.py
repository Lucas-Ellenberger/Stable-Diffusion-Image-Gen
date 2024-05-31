import csv
import os
import requests
from io import BytesIO
from PIL import Image as PILImage, UnidentifiedImageError
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from googleapiclient.discovery import build
from selenium import webdriver
import numpy as np
from PIL import Image


def fetch_image_url(query):
    try:
        # Perform a search using the Google Custom Search API
        result = service.cse().list(q=query, cx=GOOGLE_CSE_ID, searchType="image", num=1).execute()
        items = result.get("items", [])

        if items:
            # Extract the image URL from the search results
            return items[0]["link"]
        else:
            print(f"No image found for query: {query}")
            return None
    except Exception as e:
        print(f"Error fetching image URL: {e}")
        return None

def resize_and_normalize_image(file_path, target_size=(512, 512)):
    """Resize and normalize the image."""
    try:
        # Open the image
        img = PILImage.open(file_path)

        # Resize the image
        img = img.resize(target_size, PILImage.LANCZOS)


        # Normalize pixel values to range [0, 1]
        img_array = np.array(img) / 255.0

        # Create a new file path for the resized and normalized image
        resized_normalized_path = os.path.splitext(file_path)[0] + "_resized_normalized.jpg"

        # Save the resized and normalized image
        img_resized_normalized = PILImage.fromarray((img_array * 255).astype(np.uint8))
        img_resized_normalized.save(resized_normalized_path)

        return resized_normalized_path
    except Exception as e:
        print(f"Error processing image: {e}")
        return None


def download_image(image_url):
    try:
        response = requests.get(image_url, stream=True)
        if response.status_code == 200:
            # Open the image without verifying
            img = PILImage.open(BytesIO(response.content))
            return img
        else:
            print(f"Error: Failed to download image from {image_url}")
    except (UnidentifiedImageError, Exception) as e:
        print(f"Error: {e}")
    return None

# Initialize the Google Custom Search service with your API key
GOOGLE_API_KEY = 'AIzaSyAyZgmOlTt-7zgrBdrPJzSqRf6QhLzbWRA'
GOOGLE_CSE_ID = 'd5ed84b1b5451429e'
service = build("customsearch", "v1", developerKey=GOOGLE_API_KEY)

# Define the path to the CSV file
csv_file_path = 'test2prompts.csv'

# Add the directory containing the ChromeDriver executable to the PATH environment variable
chromedriver_dir = '/path/to/chromedriver_directory'
os.environ['PATH'] += os.pathsep + chromedriver_dir

# Initialize Selenium WebDriver
options = webdriver.ChromeOptions()
# options.add_argument('--headless')  # Uncomment this line to run headless Chrome
options.add_argument("--start-maximized")  # Maximize the browser window
driver = webdriver.Chrome(options=options)

# Create a new Excel workbook and select the active worksheet
workbook = Workbook()
worksheet = workbook.active

# Set the headers
worksheet.append(['Prompt', 'Image'])

# Read the CSV file
try:
    with open(csv_file_path, 'r', encoding='utf-8-sig') as csvfile:
        csvreader = csv.reader(csvfile)
        next(csvreader)  # Skip the header

        # Process each row in the CSV file
        for i, row in enumerate(csvreader, start=2):
            prompt = ','.join(row)
            print(f"Processing prompt: {prompt}")  # Print the prompt being processed
            image_url = fetch_image_url(prompt)

            if image_url:
                print(f"Image URL found: {image_url}")  # Print the image URL found
                image = download_image(image_url)
                if image:
                    print("Image downloaded successfully")  # Print if image downloaded successfully
                    # Convert image to RGB if it has an unsupported mode
                    if image.mode in ('RGBA', 'LA', 'P'):
                        image = image.convert('RGB')
                    # Save the image locally before adding it to the worksheet
                    image_path = f"downloaded_image_{i}.jpg"
                    image.save(image_path)
                    
                    # Resize and normalize the image
                    resized_image_path = resize_and_normalize_image(image_path)
                    if resized_image_path:
                        excel_image = ExcelImage(resized_image_path)
                        image_cell = f'B{i}'
                        worksheet.row_dimensions[i].height = 120  # Set row height
                        worksheet.column_dimensions['B'].width = 30  # Set column width
                        worksheet.add_image(excel_image, image_cell)
                    else:
                        print("Error: Failed to resize and normalize image")
                        worksheet.cell(row=i, column=2, value='Image not found')
                else:
                    print("Error: Image could not be downloaded")  # Print if image could not be downloaded
                    worksheet.cell(row=i, column=2, value='Image not found')
            else:
                print("Error: Image URL not found")  # Print if image URL not found
                worksheet.cell(row=i, column=2, value='Image not found')

            # Append the prompt to the worksheet
            worksheet.cell(row=i, column=1, value=prompt)

except FileNotFoundError:
    print(f"Error: CSV file '{csv_file_path}' not found.")
except Exception as e:
    print(f"Error: {e}")


# Save the workbook to a file
script_directory = os.path.dirname(os.path.realpath(__file__))
excel_file_path = os.path.join(script_directory, 'dataset_with_images.xlsx')

print("Script directory:", script_directory)

try:
    workbook.save(excel_file_path)
    print(f'Saved Excel file with images to {excel_file_path}')
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()

# Close the Selenium WebDriver
driver.quit()
